from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

class TrendyolScraper:
    def __init__(self, excel_path, logger=None):
        self.excel_path = excel_path
        self.logger = logger
        self.options = self._get_options()
        self.driver = webdriver.Edge(options=self.options)
        self.driver.maximize_window()
        self.actions = ActionChains(self.driver)
        self.sellerInfos = [
            "Mağaza Adı","Mağaza Puanı","Ürün Değerlendirme Puanı","Trendyol'daki Süre",
            "Ortalama Kargo Süresi","Değerlendirme Sayısı","Yorum Sayısı","Satıcı Değerlendirme Puanı",
            "5 Yıldızlı Değerlendirme Sayısı","4 Yıldızlı Değerlendirme Sayısı","3 Yıldızlı Değerlendirme Sayısı",
            "2 Yıldızlı Değerlendirme Sayısı","1 Yıldızlı Değerlendirme Sayısı"
        ]
        self.sellerData = pd.DataFrame(columns=self.sellerInfos)

    def log(self, message):
        if self.logger:
            self.logger(message)
        else:
            print(message)

    def _get_options(self):
        options = Options()
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--disable-web-security")
        options.add_argument("--disable-features=VizDisplayCompositor")
        options.add_argument("--disable-extensions")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--enable-javascript")
        options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0")
        options.add_experimental_option('useAutomationExtension', False)
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        return options

    def safe_find_element(self, by, identifier, timeout=15):
        try:
            return WebDriverWait(self.driver, timeout).until(EC.presence_of_element_located((by, identifier)))
        except Exception as e:
            return e

    def safe_find_elements(self, by, identifier, timeout=15):
        try:
            return WebDriverWait(self.driver, timeout).until(EC.presence_of_all_elements_located((by, identifier)))
        except Exception as e:
            return e

    def scrape_single(self, sellerText):
        self.driver.get("https://www.trendyol.com/")
        try:
            self.safe_find_element(By.CSS_SELECTOR,"button#onetrust-reject-all-handler",timeout=2).click()
        except:
            pass
        search_Box = self.safe_find_element(By.CSS_SELECTOR,'input[placeholder="Aradığınız ürün, kategori veya markayı yazınız"]')
        search_Box.clear()
        search_Box.send_keys(sellerText)
        try:
            seller = self.safe_find_element(By.CSS_SELECTOR,'div.e6w_CtSm a[href^="/magaza"]',timeout=2)
            seller.click()
        except:
            try:
                search_Box.clear()
                search_Box.send_keys(sellerText.upper())
                seller = self.safe_find_element(By.CSS_SELECTOR,'div.e6w_CtSm a[href^="/magaza"]',timeout=7)
                seller.click()
            except:
                self.log(f"⚠️ {sellerText} mağazası bulunamadı.")
                return
        try:
            self.safe_find_element(By.CSS_SELECTOR,"div.coupon-gradient a.close-button",timeout=3).click()
        except:
            pass
        try:
            sellerPoint = float(self.safe_find_element(By.CSS_SELECTOR,'div.ss-header-score',timeout=3).text)
        except:
            sellerPoint = None
        sellerProfile = self.driver.find_element(By.CSS_SELECTOR,'div.seller_profile_button')
        self.actions.click(sellerProfile).perform()
        try:
            productReviewPoint = float(self.safe_find_element(By.CSS_SELECTOR,"span.product-review-section-wrapper__wrapper__rating_wrapper_left__rating_value").text)
        except:
            productReviewPoint = None
        try:
            TrenyoldakiSure = self.safe_find_element(By.XPATH,'//*[@id="seller-profile"]/div/div/div[1]/div[1]/div/span[2]').text
        except:
            TrenyoldakiSure = "Yeterli Veri Yok"
        try:
            AverageCargoTime = self.safe_find_element(By.XPATH,'//*[@id="seller-profile"]/div/div/div[2]/div[1]/span[2]').text
        except:
            AverageCargoTime = "Yeterli Veri Yok"
        try:
            ReviewCount = int(self.safe_find_element(By.XPATH,'//*[@id="seller-profile"]/div/div/div[3]/div[2]/div/div[1]/div[2]/span[1]').text.split()[0])
        except:
            ReviewCount = 0
        try:
            CommentCount = int(self.safe_find_element(By.XPATH,'//*[@id="seller-profile"]/div/div/div[3]/div[2]/div/div[1]/div[2]/span[2]').text.split()[0])
        except:
            CommentCount = 0
        try:
            sellerReviewBtn = self.safe_find_element(By.CSS_SELECTOR,'div[data-testid="seller-review-tab"]')
            self.actions.click(sellerReviewBtn).perform()
            sellerReviewPoint = float(self.safe_find_element(By.CSS_SELECTOR,"span.seller-review-container-wrapper__wrapper__rating_wrapper__rating_value").text)
        except:
            sellerReviewPoint = None
        try:
            ratingPoints = self.safe_find_element(By.CSS_SELECTOR,'div.seller-review-container-wrapper__wrapper__rating_wrapper__arrow-down')
            self.actions.move_to_element(ratingPoints).perform()
            ratingPointDiv = self.safe_find_element(By.CSS_SELECTOR, "div.detailed-rating-modal-container__header")
            self.actions.move_to_element(ratingPointDiv).perform()
            ratings = self.safe_find_elements(By.CSS_SELECTOR,"div.detailed-rating-modal-container__detailed-rating")
            index = 2
            sellerDataRows = [sellerText,sellerPoint,productReviewPoint,TrenyoldakiSure,AverageCargoTime,ReviewCount,CommentCount,sellerReviewPoint]
            for rating in ratings:
                rate_value = self.safe_find_element(By.XPATH,f"/html/body/div[3]/div/div[2]/div/div[{index}]/span[4]").text
                sellerDataRows.append(int(rate_value))
                index += 1
            if not all(pd.isna(x) or x == "Yeterli Veri Yok" or x == 0 for x in sellerDataRows):
                self.sellerData.loc[len(self.sellerData)] = sellerDataRows
            self.log(f"✅ {sellerText} mağazası başarıyla işlendi.")
        except Exception as e:
            self.log(f"❌ {sellerText} mağazası için değerlendirme alınamadı. Hata: {e}")
            sellerDataRows = [sellerText, sellerPoint, productReviewPoint, TrenyoldakiSure, AverageCargoTime, ReviewCount, CommentCount, sellerReviewPoint,0,0,0,0,0]
            if not all(pd.isna(x) or x == "Yeterli Veri Yok" or x == 0 for x in sellerDataRows):
                self.sellerData.loc[len(self.sellerData)] = sellerDataRows

def style_excel(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
    son_satir = ws.max_row
    son_sutun = ws.max_column
    aralik = f"A1:{get_column_letter(son_sutun)}{son_satir}"
    excel_tablosu = Table(displayName="VeriTablosu", ref=aralik)
    stil = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    excel_tablosu.tableStyleInfo = stil
    ws.add_table(excel_tablosu)
    font = Font(name='Calibri', size=11, bold=False)
    baslik_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = baslik_font
        cell.alignment = alignment
        cell.fill = header_fill
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = font
            cell.alignment = alignment
    for row in range(2, ws.max_row + 1):
        ws[f'F{row}'].number_format = '0" Değerlendirme"'
        ws[f'G{row}'].number_format = '0" Yorum"'
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(excel_path)